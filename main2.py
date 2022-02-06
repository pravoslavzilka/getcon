import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import pandas as pd
import openpyxl
from os import listdir
from os.path import isfile, join
from datetime import datetime
from database import db_session
from models import Record


datime = datetime.now().strftime("%I-%M-%B-%d-%Y")
path_to_files = "input_data2/"
#writer = pd.ExcelWriter(f"output_data/{datime}.xlsx")

onlyfiles = [f for f in listdir(path_to_files) if isfile(join(path_to_files, f))]

db_session.rollback()


names = []
co_emails = []
co_numbers = []
co_azet = []

driver = webdriver.Firefox(executable_path=r'geckodriver.exe')
driver.get("https://www.azet.sk/katalog/")
driver.find_element_by_xpath('//button[contains(@class,"fc-button fc-cta-consent fc-primary-button")]').click()


for index, file in enumerate(onlyfiles):
	wb_obj = openpyxl.load_workbook(path_to_files + file)
	sheet_obj = wb_obj.active


	for i in range(1,sheet_obj.max_row):

		query = sheet_obj.cell(row=i+1, column=2).value
		pos_email = sheet_obj.cell(row=i+1, column=23).value
		town = sheet_obj.cell(row=i+1, column=6).value
		updated_town = town.split("-") if town else ""
		
		elements = []
		if pos_email:
			safe_email = pos_email.split(",")[0]
			record  = Record.query.filter(Record.email == safe_email).first()
			if not record:
				names.append(query)
				co_emails.append(safe_email + ";")
				co_azet.append("-")
				r = Record(query,safe_email + ";")
				try:
					db_session.add(r)
					db_session.commit()
				except:
					db_session.rollback()
					continue


				print(f"\n{file} Line: {i+1} -->",query, "-- excel databaáza --",safe_email,"\n")
				try:
					driver.get("https://www.azet.sk/katalog/")
				except:
					driver.close()
					driver = webdriver.Firefox(executable_path=r'geckodriver.exe')
					driver.get("https://www.azet.sk/katalog/")

			else:
				print("--- duplikát2---")
		else:
			try:
				driver.find_element_by_xpath("//input[@placeholder='Hľadaj firmy']").clear()
				driver.find_element_by_xpath("//input[@placeholder='Hľadaj firmy']").send_keys(query)

				#driver.find_element_by_xpath("//input[@placeholder='Vo všetkých mestách']").clear()
				#driver.find_element_by_xpath("//input[@placeholder='Vo všetkých mestách']").send_keys(updated_town[0])
				
				driver.find_element_by_xpath("//input[@placeholder='Hľadaj firmy']").submit()

				sleep(0.5)

				# elements = driver.find_elements(By.CSS_SELECTOR, 'div.records a')

				elements = driver.find_elements_by_css_selector("div.record a")
			except:

				continue

			if elements:
				try:
					comapany_page = elements[1].get_attribute("href")	
					driver.get(comapany_page)
				except:
					driver.get("https://www.azet.sk/katalog/")
					continue		
				try:		
					page_source = driver.page_source
				except:
					continue
					
				EMAIL_REGEX = r'''(?:[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*|"(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21\x23-\x5b\x5d-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])*")@(?:(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?|\[(?:(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9]))\.){3}(?:(2(5[0-5]|[0-4][0-9])|1[0-9][0-9]|[1-9]?[0-9])|[a-z0-9-]*[a-z0-9]:(?:[\x01-\x08\x0b\x0c\x0e-\x1f\x21-\x5a\x53-\x7f]|\\[\x01-\x09\x0b\x0c\x0e-\x7f])+)\])'''
				
				emails = []
				numbers = []
				
				for re_match in re.finditer(EMAIL_REGEX, page_source):
					emails.append(re_match.group())
				
				pattern = re.compile(r"([0-9]+( [0-9]+)+)")

				elements = driver.find_elements_by_css_selector("div.mainContact")

				for element in elements:
				    match = pattern.match(element.text)
				    if match:
				        numbers.append(element.text)

				if driver.current_url not in co_azet:         
					
					if emails:
						test_r = Record.query.filter(Record.email == emails[0]).first()
						test_r2 = Record.query.filter(Record.azet == driver.current_url).first()
						if test_r or test_r2:
							print("--- duplikát3---")	
						else:
							names.append(query)        
							co_azet.append(driver.current_url)
							co_emails.append(emails[0] + ";")
							r = Record(query, emails[0] + ";", driver.current_url)
							try:
								db_session.add(r)
								db_session.commit()
							except:
								db_session.rollback()
								continue
					else:
						continue


						
					print(f"\n{file} Line: {i+1} -->",query,driver.current_url,"\n")
					try:
						driver.get("https://www.azet.sk/katalog/")
					except:
						driver.close()
						driver = webdriver.Firefox(executable_path=r'geckodriver.exe')
						driver.get("https://www.azet.sk/katalog/")	
				else:
					print("--- duplikát ---")
			else:
				print("--- nenašlo sa ---")

		
	raw_data = {
		"Názov firmy": names,
		"Email": co_emails
	}

	df = pd.DataFrame(raw_data)
	print(df)
			
	'''df.to_excel(writer, "Sheet" + str(index))		
	writer.save()'''
	db_session.commit()
	names = []
	co_emails = []
	co_numbers = []
	co_azet = []

db_session.remove()
driver.close()

#print(driver.find_element_by_class_name("recordContainer"))


# regex pre formát +42# ### ### ### : (([+-]?(?=\.\d|\d)(?:\d+)?(?:\.?\d*))(?:[eE]([+-]?\d+))?( ([+-]?(?=\.\d|\d)(?:\d+)?(?:\.?\d*))(?:[eE]([+-]?\d+))?)+) 